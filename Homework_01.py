"""
Задание №1
а) Прочитайте из трёх excel файлов (заранее создайте их, внутри 1111, 2222, 3333).
б) Отсортируйте полученную матрицу в порядке убывания.
в) Запишите это в один файл, изменив шрифт и обернув в границы.
"""

import openpyxl
from openpyxl.styles import Font, Border, Side


def read_data_from_file(filename):
    wb = openpyxl.load_workbook(filename)
    sh = wb.active
    return sh["A1"].value


def create_excel_file(file_name, measure):
    work_book = openpyxl.Workbook()
    tables = work_book.active
    tables['A1'] = measure
    work_book.save(file_name)


create_excel_file("file1.xlsx", 1111)
create_excel_file("file2.xlsx", 2222)
create_excel_file("file3.xlsx", 3333)


file1_data = read_data_from_file("file1.xlsx")
file2_data = read_data_from_file("file2.xlsx")
file3_data = read_data_from_file("file3.xlsx")


data = [file1_data, file2_data, file3_data]
data.sort(reverse=True)


workbook = openpyxl.Workbook()
sheet = workbook.active


custom_font = Font(name='Times New Roman', size=12, bold=True)
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

for index, value in enumerate(data):
    cell = sheet.cell(row=index + 1, column=1)
    cell.value = value
    cell.font = custom_font
    cell.border = thin_border

workbook.save("output.xlsx")
workbook = openpyxl.load_workbook("output.xlsx")
